import openpyxl

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "20201127"

member_list = [
    "神山綾斗",
    "小沼蒼紫",
    "大畑武志",
    "三宅來未",
    "浜崎徳男",
    "土橋知世",
    "多田美樹子",
    "東海林仁美"
]


ws.cell(1,1).value = "No."
ws.cell(1,2).value = "氏名"

count = 0
for member in member_list:
    ws.cell(2+count,1).value = count+1
    ws.cell(2+count,2).value = member
    count += 1

wb.save("Python勉強会参加者リスト1.xlsx")
wb.close()
