import glob
import openpyxl
from openpyxl.styles import Border, Side

files = glob.glob("./fruits/*")
xlsx_files = [file for file in files if ".xlsx" in file and not "~$" in file]

RULE_LINE = Border(
    outline=True,
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style="thin", color='000000')
)

statistics_dict = {} # ディクショナリの定義

for xlsx_file in xlsx_files:
    wb = openpyxl.load_workbook(xlsx_file)

    if "統計表" in wb:
        shipment_dict = {} # ディクショナリの初期化
        ws = wb["統計表"]
        rows = ws.rows

        for row in rows:
            if row[0].value == "品目":
                item = row[1].value
            elif type(row[4].value) is int:
                shipment_dict[row[0].value] = row[4].value

        statistics_dict[item] = shipment_dict

print(statistics_dict)

new_wb = openpyxl.Workbook()
new_wb.worksheets[0].title = "果物出荷量統計"
new_wb.create_sheet(title="果物出荷量ランキング")

ws = new_wb.worksheets[0]
ws.cell(row = 1, column = 1).value = "令和元年 果物出荷量統計"
ws.cell(row = 3, column = 1).value = "品目"
ws.cell(row = 3, column = 1).border = RULE_LINE
ws.cell(row = 3, column = 2).value = "出荷量(t)"
ws.cell(row = 3, column = 2).border = RULE_LINE
count = 0
for item in statistics_dict.items():
    ws.cell(row = 4 + count, column = 1).value = item[0]
    ws.cell(row = 4 + count, column = 1).border = RULE_LINE
    ws.cell(row = 4 + count, column = 2).value = sum(item[1].values())
    ws.cell(row = 4 + count, column = 2).border = RULE_LINE
    count += 1

ws = new_wb.worksheets[1]
ws.cell(row = 1, column = 1).value = "令和元年 果物出荷量ランキング"
count = 0
rank_num = 3
for item in statistics_dict.items():
    ws.cell(row = 3 + count, column = 1).value = "品目"
    ws.cell(row = 3 + count, column = 2).value = item[0]
    ws.merge_cells(start_row = 3 + count, start_column = 2, end_row = 3 + count, end_column = 3)
    ws.cell(row = 4 + count, column = 1).value = "順位"
    ws.cell(row = 4 + count, column = 2).value = "都道府県"
    ws.cell(row = 4 + count, column = 3).value = "出荷量(t)"
    ws.cell(row = 3 + count, column = 1).border = RULE_LINE
    ws.cell(row = 3 + count, column = 2).border = RULE_LINE
    ws.cell(row = 4 + count, column = 1).border = RULE_LINE
    ws.cell(row = 4 + count, column = 2).border = RULE_LINE
    ws.cell(row = 4 + count, column = 3).border = RULE_LINE
    sorted_dict = dict(sorted(item[1].items(), key=lambda x:x[1], reverse=True))
    max_num = min(rank_num,len(sorted_dict))
    for i in range(max_num):
        ws.cell(row = 5 + count + i, column = 1).value = i + 1
        ws.cell(row = 5 + count + i, column = 2).value = list(sorted_dict.keys())[i]
        ws.cell(row = 5 + count + i, column = 3).value = list(sorted_dict.values())[i]
        ws.cell(row = 5 + count + i, column = 1).border = RULE_LINE
        ws.cell(row = 5 + count + i, column = 2).border = RULE_LINE
        ws.cell(row = 5 + count + i, column = 3).border = RULE_LINE
    count += max_num + 3


new_wb.save("令和元年_果物出荷量統計.xlsx")