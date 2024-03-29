import openpyxl

wb = openpyxl.load_workbook("名簿.xlsx")
ws = wb.worksheets[0]
roster = {} # 名簿のディクショナリ(便宜上氏名をキーとしているが本来なら社員番号等、重複しえないものをキーにすべきです)
for row in ws.rows:
    roster[row[1].value] = {
        "Number":row[0].value,
        "Mail":row[2].value
    }
wb.close()

wb = openpyxl.Workbook()
ws = wb.worksheets[0]
ws.title = "20201127"

member_list = [
    "神山綾斗",
    "小沼蒼紫",
    "大畑武志",
    "三宅來未",
    "浜崎徳男",
    "土橋知世",
    "多田美樹子",
    "東海林仁美"
]

ws.cell(1,1).value = "No."
ws.cell(1,2).value = "社員番号"
ws.cell(1,3).value = "氏名"
ws.cell(1,4).value = "メールアドレス"

count = 0
for member in member_list:
    ws.cell(2+count,1).value = count+1
    ws.cell(2+count,2).value = roster[member]["Number"]
    ws.cell(2+count,3).value = member
    ws.cell(2+count,4).value = roster[member]["Mail"]
    count += 1

wb.save("Python勉強会参加者リスト2.xlsx")
wb.close()